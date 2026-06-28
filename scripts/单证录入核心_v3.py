#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
单证录入核心脚本 v3 — 直接 XML 操作（绕过 openpyxl）
================================================================
v2 之前用 openpyxl 加载 .xlsm 模板再保存，导致：
  - vbaProject.bin 丢失（50KB VBA 宏被丢弃）
  - sheet 1/2/3 被压缩（编号重排）
  - 数据验证扩展（dataValidation）被移除
  - sharedStrings.xml、customXml 等丢失
  → Excel 弹出"修复"对话框 → 用户体感"无法打开"

v3 方案：
  1. 复制模板 zip 整体（保留 vbaProject.bin / 所有 sheet）
  2. 只修改需要填充的 4 个 sheet 的 XML（按字符串插入 <row>）
  3. 重新打包 zip 输出

用法:
    python3 scripts/单证录入核心_v3.py
    python3 scripts/单证录入核心_v3.py <crew.xlsx> <port.xlsx> [output.xlsm]
"""

import os
import re
import sys
import json
import random
import shutil
import zipfile
import io
from datetime import datetime


WORKDIR = "/Users/qianlingbo/单证录入工作区"
INPUT_DIR = os.path.join(WORKDIR, "input")
OUTPUT_DIR = os.path.join(WORKDIR, "output")
TEMPLATES_DIR = os.path.join(WORKDIR, "templates")
SCRIPTS_DIR = os.path.join(WORKDIR, "scripts")

DEFAULT_CREW = os.path.join(INPUT_DIR, "CREW_LIST.xlsx")
DEFAULT_POC = os.path.join(INPUT_DIR, "PORT_OF_CALL.xlsx")
TEMPLATE_XLSM = os.path.join(TEMPLATES_DIR, "标准格式.xlsm")
DEFAULT_OUTPUT = os.path.join(OUTPUT_DIR, "单证录入标准格式.xlsm")


# ============================================================
# 字段映射（与 v2 相同）
# ============================================================

NAT_MAP = {
    "CHINESE": "中国", "CHINA": "中国", "CN": "中国",
    "VIETNAMESE": "越南", "VIETNAM": "越南", "VN": "越南",
    "MYANMAR": "缅甸", "BURMESE": "缅甸", "MM": "缅甸",
    "INDONESIAN": "印度尼西亚", "INDONESIA": "印度尼西亚", "ID": "印度尼西亚",
    "PHILIPPINE": "菲律宾", "PHILIPPINES": "菲律宾", "PHILIPPINO": "菲律宾", "FILIPINO": "菲律宾", "PH": "菲律宾",
    "PANAMANIAN": "巴拿马", "PANAMA": "巴拿马", "PA": "巴拿马",
    "INDIAN": "印度", "INDIA": "印度", "IN": "印度",
}

NAT_CODE = {
    "中国": "CN", "越南": "VN", "缅甸": "MM", "印度尼西亚": "ID",
    "菲律宾": "PH", "巴拿马": "PA", "印度": "IN",
}

DUTY_CODE = {
    "MASTER": "51", "CAPT": "51", "CAPTAIN": "51",
    "C/O": "52", "C.O.": "52", "CH. OFF": "52", "CHIEF OFFICER": "52", "CH OFF": "52",
    "2/O": "53", "2ND OFFICER": "53", "2ND OFF": "53", "2ND OFF.": "53",
    "3/O": "54", "3RD OFFICER": "54", "3RD OFF": "54",
    "OS": "55", "OLR": "55", "ORDINARY SEAMAN": "55", "C/CK": "55", "CH. COOK": "55", "CH COOK": "55",
    "AB": "55", "ABLE SEAMAN": "55", "AB1": "55", "AB2": "55", "AB3": "55",
    "BOSUN": "56", "BSN": "56", "BOATSWAIN": "56",
    "C/E": "61", "CHIEF ENGINEER": "61", "CH/ENGR": "61", "CH.ENGR": "61",
    "2/E": "63", "2ND ENGINEER": "63", "2ND ENGR": "63", "2ND ENGR.": "63",
    "3/E": "64", "3RD ENGINEER": "64", "3RD ENGR": "64", "3RD ENGR.": "64",
    "4/E": "65", "4TH ENGINEER": "65", "4TH ENGR": "65", "4TH ENGR.": "65",
    "ETR": "65", "FTR": "65", "FITTER": "65", "OILER": "65",
    "ELECTRICIAN": "65", "ELECT": "65", "OIL1": "65", "OIL2": "65", "OILER1": "65", "OILER2": "65",
    "WIPER": "65",
    "MESSMAN": "65",
    "OS1": "55", "OS2": "55",  # 普通水手
}

PORT_CODE = {
    # 中国
    "ZHOUSHAN": "CNZOS", "舟山": "CNZOS",
    "ZHANGJIAGANG": "CNZJG", "张家港": "CNZJG",
    "CHANGZHOU": "CNCZX", "常州": "CNCZX",
    "TIANJIN": "CNTXG", "天津": "CNTXG",
    "LANSHAN": "CNLSN", "岚山": "CNLSN",
    "BAYUQUAN": "CNBYQ", "鲅鱼圈": "CNBYQ",
    "CAOFEIDIAN": "CNCFD", "曹妃甸": "CNCFD",
    "LIANYUNGANG": "CNLYG", "连云港": "CNLYG",
    "NINGDE": "CNNDS", "宁德": "CNNDS",
    "SHANGHAI": "CNSHA", "上海": "CNSHA",
    "NINGBO": "CNNGB", "宁波": "CNNGB",
    "QINGDAO": "CNTAO", "青岛": "CNTAO",
    "RIZHAO": "CNRZH", "日照": "CNRZH",
    "DONGGUAN": "CNDGG", "东莞": "CNDGG",
    "GUANGZHOU": "CNGZG", "GUANGDONG": "CNGZG", "广州": "CNGZG", "广东": "CNGZG",
    "LUOYUAN": "CNLYA", "罗源": "CNLYA",
    # 印尼
    "MOROWALI": "IDMOR",
    "POMALAA": "IDPUM",
    "OBI ISLAND": "IDOBI", "OBI": "IDOBI",
    "MUARA BERAU": "IDSRI", "SUNGAI BERAU": "IDSRI", "IDSRI": "IDSRI",
    "TANJUNG BARA": "IDTBA", "IDTBA": "IDTBA",
    # 菲律宾
    "SURIGAO": "PHSUG", "PHSUG": "PHSUG",
    "SAN FERNANDO CEBU": "PHCEB", "CEBU": "PHCEB", "PHCEB": "PHCEB",
    # 文莱
    "MUARA": "BNMUA", "BNMUA": "BNMUA",
    # 新加坡
    "SINGAPORE": "SGSIN", "SGSIN": "SGSIN",
    # 香港
    "HONG KONG": "HKHKG", "HONGKONG": "HKHKG", "HKHKG": "HKHKG",
    # 公海
    "OPEN SEA": "OPSEA", "OPENSEA": "OPSEA", "OPSEA": "OPSEA",
    # 国际
    "HITACHINAKA": "JPHIC",
    "GUANGDONG, CHINA": "CNGZG",
    "CEBU, PHILIPPINES": "PHCEB",
    "SURIGAO, PHILIPPINES": "PHSUG",
    "DAVAO, PHILIPPINES": "PHDVO",
    "GUANGDONG": "CNGZG",
}


# ============================================================
# 解析输入（与 v2 相同）
# ============================================================

def normalize_date(s):
    """宽松日期解析 - 支持 英文月份 (JAN/Feb) 和 中文 (2026年7月)"""
    if s is None: return None
    if isinstance(s, datetime): return s
    s = str(s).strip()
    s = re.sub(r'\s+', ' ', s)
    s = s.replace(' :', ':').replace(': ', ':')
    s = re.sub(r'(\d{4})/(\d{1,2})/ ?(\d{1,2})', r'\1/\2/\3', s)

    # 英文月份映射
    en_months = {
        'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
        'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
        'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12',
    }
    # 把 "29 JAN 1969" 转为 "29 01 1969" / "31 Mar 2026" -> "31 03 2026"
    en_pattern = re.compile(r'\b(\d{1,2})\s+([A-Za-z]{3})\s+(\d{2,4})\b')
    m = en_pattern.search(s)
    if m:
        d, mon, y = m.groups()
        if mon.upper() in en_months:
            s = s.replace(m.group(0), f'{d}/{en_months[mon.upper()]}/{y}')

    # "Aug 12, 2029" -> "08/12/2029" / "Jul 05, 2029" -> "07/05/2029"
    en_pattern2 = re.compile(r'\b([A-Za-z]{3})\s+(\d{1,2}),?\s+(\d{2,4})\b')
    m = en_pattern2.search(s)
    if m:
        mon, d, y = m.groups()
        if mon.upper() in en_months:
            s = s.replace(m.group(0), f'{en_months[mon.upper()]}/{d}/{y}')

    fmts = [
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d", "%Y/%m/%d", "%Y%m%d",
        "%d/%m/%Y", "%m/%d/%Y", "%d %m %Y",
    ]
    for f in fmts:
        try: return datetime.strptime(s, f)
        except ValueError: continue
    return None


def parse_crew_xlsx(path):
    """解析 IMO Crew List - 支持两种格式:
    1) 旧格式: 表头有 FAMILY NAME 列
    2) 新格式 (IMCO FAL Form 1969): R10 是表头, 字段用 7. No. 8. Family name 等
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["ARR"] if "ARR" in wb.sheetnames else wb[wb.sheetnames[0]]

    # 找标题行 - 看是不是新格式
    is_new_format = False
    header_row = None
    for r in range(1, 15):
        cells = [str(c.value or '') for c in ws[r][:15]]
        cell_text = ' '.join(cells)
        # 新格式: R10 含 "Family name" 和 "Rank" 和 "Nationality"
        if 'FAMILY NAME' in cell_text.upper() and 'RANK' in cell_text.upper():
            header_row = r
            is_new_format = True
            break
        # 旧格式: 直接 FAMILY NAME
        if 'FAMILY NAME' in [c.upper() for c in cells]:
            header_row = r
            break
    if header_row is None: header_row = 7

    # 找列索引
    headers = [str(ws.cell(header_row, c).value or '').upper() for c in range(1, 20)]

    if is_new_format:
        # 新格式: 找 "8. FAMILY NAME" / "9. RANK" 等
        col = {}
        for i, h in enumerate(headers):
            hu = h.upper()
            if 'NO.' in hu and 'no' not in col: col['no'] = i + 1
            elif 'FAMILY NAME' in hu: col['name'] = i + 1
            elif 'GENDER' in hu or 'SEX' in hu: col['sex'] = i + 1
            elif 'RANK' in hu: col['rank'] = i + 1
            elif 'NATIONALITY' in hu and 'nat' not in col: col['nat'] = i + 1
            elif 'BIRTH' in hu: col['birth'] = i + 1
            elif 'PASSPORT' in hu: col['passport'] = i + 1
            elif 'SEAMAN' in hu: col['seaman'] = i + 1
            elif 'SIGN ON' in hu: col['signedon'] = i + 1
            elif 'PLACE' in hu and 'sign' not in col: col['place'] = i + 1

        # 备用默认值 (新格式常见布局)
        col.setdefault('no', 2)
        col.setdefault('name', 3)
        col.setdefault('rank', 4)
        col.setdefault('sex', 5)
        col.setdefault('nat', 6)
        col.setdefault('birth', 7)
        col.setdefault('passport', 8)
        col.setdefault('seaman', 9)
        col.setdefault('signedon', 10)
    else:
        col = {}
        for kw, idx_name in [('NO.', 'no'), ('NAME', 'name'), ('SEX', 'sex'),
                              ('RANK', 'rank'), ('BIRTH', 'birth'), ('NATIONAL', 'nat'),
                              ('PLACE', 'place'), ('SEAMAN', 'seaman'), ('PASSPORT', 'passport'),
                              ('SIGNED ON', 'signedon')]:
            for i, h in enumerate(headers):
                if kw in h and idx_name not in col:
                    col[idx_name] = i + 1

    crews = []
    for r in range(header_row + 1, ws.max_row + 1):
        no = ws.cell(r, col.get('no', 1)).value
        if no is None: continue
        # 跳过中文行 / #REF! 行
        nm = str(ws.cell(r, col.get('name', 2)).value or '')
        if not re.search(r'[A-Z]', nm): continue
        if '#REF' in nm or '#REF' in str(ws.cell(r, col.get('rank', 4)).value or ''): continue
        try:
            no_int = int(float(str(no).strip()))
        except: continue

        rank = str(ws.cell(r, col.get('rank', 5)).value or '').strip().upper()
        nat = str(ws.cell(r, col.get('nat', 6)).value or '').strip().upper()
        # 出生地: 城市 + 日期 (可能含 \n)
        bd_cell = ws.cell(r, col.get('birth', 5)).value
        if bd_cell:
            # 可能是 "TANZA CAVITE \n 29 JAN 1969" - 拆分找日期
            bd_str = str(bd_cell)
            if '\n' in bd_str:
                # 找包含月份名的行
                parts = bd_str.split('\n')
                for p in parts:
                    p = p.strip()
                    if re.search(r'[A-Za-z]{3}', p) and re.search(r'\d', p):
                        bd = normalize_date(p)
                        break
            else:
                bd = normalize_date(bd_str)
        else:
            bd = None

        # 护照号 / 海员证号: 新格式中是 "P7104991A\n 09 MAY 2028" - 取第一行
        passport_raw = str(ws.cell(r, col.get('passport', 8)).value or '').strip()
        seaman_raw = str(ws.cell(r, col.get('seaman', 9)).value or '').strip()
        passport = passport_raw.split('\n')[0].strip() if passport_raw else ''
        seaman = seaman_raw.split('\n')[0].strip() if seaman_raw else ''

        # signed on: 可能是 "GUANGDONG, CHINA\n 31 Mar 2026"
        so_cell = ws.cell(r, col.get('signedon', 10)).value
        so_date = None
        so_place = ''
        if so_cell:
            so_str = str(so_cell)
            # 拆 city / date
            parts = so_str.split('\n')
            so_place = parts[0].strip() if parts else ''
            if len(parts) > 1:
                so_date = normalize_date(parts[1])
            else:
                so_date = normalize_date(so_cell)

        crews.append({
            'no': no_int,
            'name_en': nm.strip(),
            'sex': '1' if 'M' in str(ws.cell(r, col.get('sex', 3)).value or '').upper() else '2',
            'rank': rank,
            'birth': bd,
            'nat_en': nat,
            'seaman_no': seaman,
            'passport_no': passport,
            'signon_date': so_date,
            'signon_place': so_place,
        })
    return crews


def parse_poc_xlsx(path):
    """解析 Port of Call - 支持 xlsx 和 PDF"""
    if path.lower().endswith('.pdf'):
        return parse_poc_pdf(path)
    return parse_poc_xlsx_native(path)


def parse_poc_pdf(path):
    """从 PDF 提取港口列表

    PDF 格式: VESSEL NAME / FLAG / 然后是表格 NO. | NAME | UNCODE | VOY | MARSEC | ARRIVED | SAILED | REMARKS
    """
    import re
    import fitz  # PyMuPDF
    doc = fitz.open(path)
    full_text = ''
    for page in doc:
        full_text += page.get_text() + '\n'

    # 提取头部信息
    vessel = ''
    flag = ''
    port_of_arrival = ''
    m = re.search(r'VESSEL NAME:\s*(.+)', full_text)
    if m: vessel = m.group(1).strip()
    m = re.search(r'FLAG:\s*(.+)', full_text)
    if m: flag = m.group(1).strip()
    m = re.search(r'PORT OF ARRIVAL:\s*(.+)', full_text)
    if m: port_of_arrival = m.group(1).strip()

    # 找表格行: NO. NUMBER | NAME | UNCODE | VOY | MARSEC | ARRIVED | SAILED | REMARKS
    # 行格式: NO. \n NAME (可能有逗号后的国家) \n UNCODE \n VOY \n MARSEC \n ARRIVED \n SAILED \n REMARKS
    records = []

    # 简化策略: 用 "NAME OF PORTS" 作为表头, 找后续行
    # PDF 中每条港口记录是连续 8-9 行
    lines = full_text.split('\n')
    i = 0
    # 找到表头之后开始
    while i < len(lines):
        if 'NAME OF PORTS' in lines[i]:
            i += 1
            break
        i += 1
    # 跳过表头
    while i < len(lines) and 'REMARKS' not in lines[i]:
        i += 1
    i += 1  # 跳过 REMARKS 行

    while i < len(lines):
        line = lines[i].strip()
        # 找行号行 (纯数字 1-10)
        if line.isdigit() and 1 <= int(line) <= 20:
            no = int(line)
            i += 1
            # 读 NAME (可能跨多行: "SAN FERNANDO CEBU, \n PHILIPPINES")
            name = lines[i].strip()
            i += 1
            # 如果下一行不是 UNLOCODE (5字母代码), 是国家名 — 把它加到 name
            # UNLOCODE 特征: 5 个大写字母
            if i < len(lines) and not re.match(r'^[A-Z]{5}$', lines[i].strip()):
                name = name + ' ' + lines[i].strip()
                i += 1

            uncode = lines[i].strip() if i < len(lines) else ''
            i += 1
            voy = lines[i].strip() if i < len(lines) else ''
            i += 1
            marsec = lines[i].strip() if i < len(lines) else ''
            i += 1
            arr_str = lines[i].strip() if i < len(lines) else ''
            i += 1
            dep_str = lines[i].strip() if i < len(lines) else ''
            i += 1
            # REMARKS 可能是单行也可能是多行 (如 "LOADING  COAL" 或 "DISCHARGING  CLINKER")
            # 简单处理: 一直读到下一行是数字 (下一个港口号)
            remarks_parts = []
            while i < len(lines):
                l = lines[i].strip()
                if l.isdigit() and 1 <= int(l) <= 20:
                    break
                if 'CAPT.' in l or 'MASTER' in l or 'PORT OF' in l or 'DATE OF' in l or 'LAST' in l.upper() or l == '':
                    break
                remarks_parts.append(l)
                i += 1
            remarks = ' '.join(remarks_parts).strip()

            # 转换日期
            arr_date = parse_pdf_date(arr_str) if arr_str else None
            dep_date = parse_pdf_date(dep_str) if dep_str else None

            # 拆分国家/港口
            # PDF 中 NAME 字段是 "PORT_NAME, COUNTRY" 格式
            # 例: "MUARA BERAU, INDONESIA" -> port="MUARA BERAU" country="INDONESIA"
            name_upper = name.upper().strip()
            country = ''
            port = name_upper

            # 1) 找 ", COUNTRY" 形式 - 优先匹配最长国家名
            for cn in ['HONG KONG, CHINA', 'INDONESIA', 'PHILIPPINES', 'SINGAPORE', 'CHINA', 'BRUNEI']:
                if (', ' + cn) in name_upper:
                    port = name_upper.split(', ' + cn)[0].strip()
                    country = cn
                    break
            # 2) 特殊情况: "HONG KONG" 没有逗号
            if not country and 'HONG KONG' in name_upper:
                country = 'HONG KONG'
                port = 'HONG KONG'

            country_code = uncode[:2] if uncode else ''

            records.append({
                'no': no,
                'port': port,
                'country': country,
                'uncode': uncode,
                'arr': arr_date,
                'dep': dep_date,
                'remarks': remarks,
                'marsec': marsec,
            })
            continue
        i += 1

    return records


def parse_pdf_date(s):
    """解析 PDF 日期: 24-Jun-26 -> datetime"""
    import re
    from datetime import datetime
    s = s.strip()
    if not s: return None
    months = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
              'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12}
    m = re.match(r'(\d{1,2})[-\s](\w{3})[-\s](\d{2,4})', s)
    if not m: return None
    day, mon, year = m.groups()
    year = int(year)
    if year < 100: year += 2000
    return datetime(year, months.get(mon.upper(), 1), int(day))


def parse_poc_xlsx_native(path):
    """解析 xlsx 格式的 Port of Call"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    # 找表头行
    header_row = None
    for r in range(1, 10):
        cells = [str(c.value or '').upper() for c in ws[r][:10]]
        if 'PORT' in cells and 'COUNTRY' in cells:
            header_row = r
            break
    if header_row is None: header_row = 3

    headers = [str(ws.cell(header_row, c).value or '').upper() for c in range(1, 12)]
    col = {}
    for i, h in enumerate(headers):
        hu = h.upper()
        if 'NO.' in hu and 'no' not in col: col['no'] = i + 1
        elif 'PURPOSE' in hu and 'purpose' not in col: col['purpose'] = i + 1
        elif hu == 'PORT' and 'port' not in col: col['port'] = i + 1
        elif 'COUNTRY' in hu and 'country' not in col: col['country'] = i + 1
        elif 'ARRIVAL' in hu and 'arr' not in col: col['arr'] = i + 1
        elif 'DEPARTURE' in hu and 'dep' not in col: col['dep'] = i + 1
        elif 'SHIP SECURITY' in hu and 'ssl' not in col: col['ssl'] = i + 1
        elif 'PORT' in hu and 'SECURITY' in hu and 'psl' not in col: col['psl'] = i + 1

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        no = ws.cell(r, col.get('no', 1)).value
        port = ws.cell(r, col.get('port', 3)).value
        if no is None or port is None: continue
        try:
            no_int = int(float(str(no).strip()))
        except: continue
        records.append({
            'no': no_int,
            'purpose': str(ws.cell(r, col.get('purpose', 2)).value or '').strip(),
            'port': str(port).strip(),
            'country': str(ws.cell(r, col.get('country', 4)).value or '').strip().upper(),
            'arr': normalize_date(ws.cell(r, col.get('arr', 5)).value),
            'dep': normalize_date(ws.cell(r, col.get('dep', 6)).value),
            'ssl': str(ws.cell(r, col.get('ssl', 7)).value or '1').strip(),
            'psl': str(ws.cell(r, col.get('psl', 8)).value or '1').strip(),
        })
    # 序号重排
    for i, rec in enumerate(records, 1):
        rec['no'] = i
    return records


# ============================================================
# XML 工具
# ============================================================

def cell_ref(col_letter, row):
    return f"{col_letter}{row}"


def escape_xml(s):
    if s is None: return ''
    s = str(s)
    return (s.replace('&', '&amp;')
             .replace('<', '&lt;')
             .replace('>', '&gt;')
             .replace('"', '&quot;')
             .replace("'", '&apos;'))


def build_row(row_num, values, start_col='A', style=None):
    """生成 <row> XML 字符串
    
    values: list of (col_letter, value, [type])
      type: 's' sharedString, 'str' inline string, 'n' number, default str
    """
    cells_xml = []
    for item in values:
        col = item[0]
        val = item[1]
        vtype = item[2] if len(item) > 2 else 'str'
        ref = f"{col}{row_num}"
        if vtype == 'n':
            cells_xml.append(f'<c r="{ref}"><v>{val}</v></c>')
        elif vtype == 's':
            cells_xml.append(f'<c r="{ref}" t="s"><v>{val}</v></c>')
        else:
            cells_xml.append(f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{escape_xml(val)}</t></is></c>')
    style_attr = f' s="{style}"' if style else ''
    return f'<row r="{row_num}"{style_attr}>' + ''.join(cells_xml) + '</row>'


def insert_rows_into_sheet_xml(sheet_xml, new_rows_xml, after_row):
    """在指定行号之后插入新行 XML"""
    # 找到 </row> 在 after_row 后面的位置
    pattern = rf'(<row r="{after_row}"[^>]*>.*?</row>)'
    match = re.search(pattern, sheet_xml, re.DOTALL)
    if not match:
        # 找不到精确行 — 尝试找所有 <row> 然后定位
        all_rows = list(re.finditer(r'<row r="(\d+)"', sheet_xml))
        if not all_rows:
            return sheet_xml.replace('</sheetData>', '<sheetData>' + ''.join(new_rows_xml) + '</sheetData>')
        # 找位置
        target_idx = 0
        for i, m in enumerate(all_rows):
            if int(m.group(1)) > after_row:
                target_idx = i
                break
            target_idx = i + 1
        if target_idx == 0:
            insert_pos = all_rows[0].start()
        else:
            # 在 target_idx 之前插入（即在它前面）
            insert_pos = all_rows[target_idx].start()
    else:
        insert_pos = match.end()
    return sheet_xml[:insert_pos] + ''.join(new_rows_xml) + sheet_xml[insert_pos:]


# ============================================================
# 数据准备
# ============================================================

def build_crew_rows(crews, shared_strings):
    """生成船员名单行（船上非旅客人员清单）"""
    rows = []
    for i, c in enumerate(crews, 3):  # 从 R3 开始
        # 中文名（如果有）或英文
        name = c['name_en']  # 简化: v3 暂只填英文
        nat_cn = NAT_MAP.get(c['nat_en'].upper(), c['nat_en'])
        nat_code = NAT_CODE.get(nat_cn, '')
        rank = DUTY_CODE.get(c['rank'].upper(), '55')
        is_cn = (nat_cn == '中国')
        cert_type = '17-海员证' if is_cn else '14-普通护照'
        cert_no = c['seaman_no'] if is_cn else c['passport_no']
        birth_str = c['birth'].strftime('%Y%m%d') if c['birth'] else ''
        signon_str = c['signon_date'].strftime('%Y%m%d') if c['signon_date'] else ''
        signon_port = PORT_CODE.get(c['signon_place'].upper(), '')

        row = build_row(i, [
            ('A', i - 2),                # 序号
            ('B', name),                 # 姓名
            ('C', f"{c['sex']}-{'男' if c['sex']=='1' else '女'}"),  # 性别
            ('D', f"{rank}-{rank_name(rank)}"),  # 职务
            ('E', f"{nat_code}-{nat_cn}"),  # 国籍
            ('F', birth_str),            # 出生日期
            ('G', nat_cn),               # 出生地点
            ('H', cert_type),            # 证件类型
            ('I', cert_no),              # 证件号码
            ('M', signon_str),           # 登船日期
            ('N', signon_port),          # 登船口岸
        ])
        rows.append(row)
    return rows


def rank_name(code):
    return {
        '51': '船长', '52': '大副', '53': '二副', '54': '三副',
        '55': '值班水手', '56': '高级值班水手',
        '61': '轮机长', '62': '大管轮', '63': '二管轮', '64': '三管轮',
        '65': '值班机工', '66': '高级值班机工',
    }.get(code, '值班水手')


def build_goods_rows(crews):
    rows = []
    for i, c in enumerate(crews, 3):
        nat_cn = NAT_MAP.get(c['nat_en'].upper(), c['nat_en'])
        is_cn = (nat_cn == '中国')
        cert_type = '17-海员证' if is_cn else '14-普通护照'
        cert_no = c['seaman_no'] if is_cn else c['passport_no']
        row = build_row(i, [
            ('A', i - 2),
            ('B', cert_type),
            ('C', cert_no),
            ('D', '0100'),
            ('E', '计算机'),
            ('F', '1'),
            ('G', '001'),
        ])
        rows.append(row)
    return rows


def build_port_call_rows(records):
    """海事船岸活动信息 (sheet: 海事船岸活动信息)"""
    rows = []
    for i, r in enumerate(records, 3):
        # 进港 0-11, 离港 12-23
        if r['arr']:
            arr_h = random.randint(0, 11)
            arr_m = random.randint(0, 59)
            arr_s = random.randint(0, 59)
            arr_str = r['arr'].replace(hour=arr_h, minute=arr_m, second=arr_s).strftime('%Y/%m/%d %H:%M:%S')
        else:
            arr_str = ''
        if r['dep']:
            dep_h = random.randint(12, 23)
            dep_m = random.randint(0, 59)
            dep_s = random.randint(0, 59)
            dep_str = r['dep'].replace(hour=dep_h, minute=dep_m, second=dep_s).strftime('%Y/%m/%d %H:%M:%S')
        else:
            dep_str = ''
        port_code = PORT_CODE.get(r['port'].upper(), '')

        # 国家代码 - 从 uncode 取前 2 字母
        country = r['country'].upper()
        if country == 'UN' or 'OPEN' in r['port'].upper():
            country = 'UN'
        elif country in ('CHINA', 'CN'): country = 'CN'
        elif country in ('INDONESIA', 'ID'): country = 'ID'
        elif country in ('PHILIPPINES', 'PH'): country = 'PH'
        elif country in ('SINGAPORE', 'SG'): country = 'SG'
        elif country in ('HONG KONG', 'HONG KONG, CHINA', 'HK'): country = 'HK'
        elif country in ('BRUNEI', 'BN'): country = 'BN'
        else:
            # 用 uncode 前 2 字母 (IDSRI->ID, SGSIN->SG, PHCEB->PH)
            if r.get('uncode'):
                country = r['uncode'][:2].upper()
            else:
                country = 'UN'

        row = build_row(i, [
            ('A', i - 2),
            ('B', arr_str),
            ('C', dep_str),
            ('D', country),
            ('E', '1-1级'),
            ('G', port_code),
            ('H', '1-1级'),
        ])
        rows.append(row)
    return rows


def build_top10_rows(records):
    """前十港信息"""
    rows = []
    for i, r in enumerate(records[:10], 3):
        if r['arr']:
            arr_h = random.randint(0, 11)
            arr_m = random.randint(0, 59)
            arr_s = random.randint(0, 59)
            arr_str = r['arr'].replace(hour=arr_h, minute=arr_m, second=arr_s).strftime('%Y/%m/%d %H:%M:%S')
        else:
            arr_str = ''
        if r['dep']:
            dep_h = random.randint(12, 23)
            dep_m = random.randint(0, 59)
            dep_s = random.randint(0, 59)
            dep_str = r['dep'].replace(hour=dep_h, minute=dep_m, second=dep_s).strftime('%Y/%m/%d %H:%M:%S')
        else:
            dep_str = ''
        port_code = PORT_CODE.get(r['port'].upper(), '')
        # 前十港用国家全名 (与海事船岸用代码不同)
        country = r['country']
        if not country and r.get('uncode'):
            # 从 uncode 前2字母反推国家名
            cc = r['uncode'][:2].upper()
            country = {
                'CN': 'CHINA', 'ID': 'INDONESIA', 'PH': 'PHILIPPINES',
                'SG': 'SINGAPORE', 'HK': 'HONG KONG', 'BN': 'BRUNEI',
                'MY': 'MALAYSIA', 'JP': 'JAPAN', 'KR': 'KOREA',
            }.get(cc, '')
        if country == 'UN' or 'OPEN' in r['port'].upper():
            country = 'UN'

        row = build_row(i, [
            ('A', i - 2),
            ('B', arr_str),
            ('C', dep_str),
            ('E', country),
            ('F', port_code),
        ])
        rows.append(row)
    return rows


# ============================================================
# 主流程：复制模板 zip + 修改 4 个 sheet XML
# ============================================================

# 模板里的 sheet 编号映射（来自原模板的 workbook.xml 顺序）
SHEET_RID_TO_PATH = {
    'rId1': ('船上非旅客人员清单', 'xl/worksheets/sheet1.xml'),
    'rId2': ('旅客清单', 'xl/worksheets/sheet2.xml'),
    'rId3': ('供退物料清单', 'xl/worksheets/sheet3.xml'),
    'rId4': ('船上非旅客人员物品清单', 'xl/worksheets/sheet4.xml'),
    'rId5': ('船用物品清单', 'xl/worksheets/sheet5.xml'),
    'rId6': ('前十港信息', 'xl/worksheets/sheet6.xml'),
    'rId7': ('危险品信息', 'xl/worksheets/sheet7.xml'),
    'rId8': ('船舶证书信息', 'xl/worksheets/sheet8.xml'),
    'rId9': ('压舱水详细信息', 'xl/worksheets/sheet9.xml'),
    'rId10': ('海事船岸活动信息', 'xl/worksheets/sheet10.xml'),
    'rId11': ('沿海空箱信息', 'xl/worksheets/sheet11.xml'),
    'rId12': ('压舱水报告单信息', 'xl/worksheets/sheet12.xml'),
    'rId13': ('压舱水装载信息', 'xl/worksheets/sheet13.xml'),
    'rId14': ('压舱水更换信息表信息', 'xl/worksheets/sheet14.xml'),
    'rId15': ('压舱水排放信息表信息', 'xl/worksheets/sheet15.xml'),
    'rId16': ('随船人员清单', 'xl/worksheets/sheet16.xml'),
    'rId17': ('参数', 'xl/worksheets/sheet17.xml'),
}


def find_data_row_range(sheet_xml):
    """找 sheetData 中 R3 起的数据占位行范围

    模板里每个可填写 sheet 的 R1 是表头, R2 是说明, R3-Rn 是空数据行（带 styles），
    我们要替换 R3-Rn 段。

    返回: (start, end) — 替换的起止行号, end 之后保留
    """
    # 找 dimension 范围, 跳过 R1/R2
    dim = re.search(r'<dimension ref="([A-Z]+)(\d+):([A-Z]+)?(\d+)"', sheet_xml)
    if dim:
        # 模板结构: R1=表头, R2=说明, R3-Rn=数据占位
        # dimension 给的 R1 是表头, 所以数据从 R3 开始
        return 3, int(dim.group(4))
    # fallback
    rows = re.findall(r'<row r="(\d+)"', sheet_xml)
    if len(rows) >= 3:
        return 3, max(int(r) for r in rows)
    return 3, 3


def fill_sheet_xml(sheet_xml, new_rows_xml):
    """替换 sheetData 中 R3 起的占位行为新数据

    关键: 不能追加在 </sheetData> 前 (会行号冲突)
    必须用正则把 R3-Rn 的占位行整段删除, 然后在 R2 后插入新行
    """
    if not new_rows_xml:
        return sheet_xml

    start, end = find_data_row_range(sheet_xml)

    # 删除 R{start} 到 R{end} 之间的所有 <row>...</row>
    # 用一个 broad 正则: 从 R{start} 开始, 一路匹配到 R{end} 之后的第一个非 row 元素
    # 更安全: 一个个删除
    for r in range(start, end + 1):
        sheet_xml = re.sub(
            rf'<row r="{r}"[^>]*>.*?</row>',
            '',
            sheet_xml,
            count=1,
            flags=re.DOTALL,
        )

    # 在 R{start-1} (即 R2) 的 </row> 之后插入新行
    # 找 R{start-1} 的 </row> 位置
    prev = start - 1
    m = re.search(rf'<row r="{prev}"[^>]*>.*?</row>', sheet_xml, re.DOTALL)
    if m:
        insert_pos = m.end()
    else:
        # fallback: 第一个 <row 之后
        m2 = re.search(r'<row r="\d+"', sheet_xml)
        insert_pos = m2.end() if m2 else 0

    # 同时更新 dimension 范围 - 用精确正则避免贪婪
    if new_rows_xml:
        last_row = start + len(new_rows_xml) - 1
        # 匹配 <dimension ref="A1:Xn" /> (Xn 是任意数字, X 是任意列字母)
        sheet_xml = re.sub(
            r'(<dimension ref="[A-Z]+\d+:)([A-Z]+)(\d+)(")',
            lambda m: m.group(1) + m.group(2) + str(last_row) + m.group(4),
            sheet_xml,
            count=1,
        )

    return sheet_xml[:insert_pos] + ''.join(new_rows_xml) + sheet_xml[insert_pos:]


def process(template_path, output_path, crews, records, output_format='xlsx'):
    """核心：复制模板 zip + 修改 4 个 sheet XML

    output_format: 'xlsx' (无宏) 或 'xlsm' (保留宏)
    'xlsx' 推荐 — Excel 不会弹"文件格式不匹配"警告
    'xlsm' 保留 vbaProject.bin + 数据验证下拉, 但 Excel 会弹警告
    """
    modified_sheets = {}

    if output_format == 'xlsx':
        # xlsx 模式: 把 .xlsm 模板改造成 .xlsx (去掉 vbaProject 引用)
        # 但保留 dataValidations 扩展 (openpyxl 会丢)
        # 简单做法: 直接复制原模板, 改 [Content_Types] 和 workbook.xml.rels
        pass

    # 1. 船员名单 → sheet1
    crew_rows = build_crew_rows(crews, None)
    if crew_rows:
        with zipfile.ZipFile(template_path) as z:
            sheet1 = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
        modified_sheets['xl/worksheets/sheet1.xml'] = fill_sheet_xml(sheet1, crew_rows)

    # 2. 物品清单 → sheet4
    goods_rows = build_goods_rows(crews)
    if goods_rows:
        with zipfile.ZipFile(template_path) as z:
            sheet4 = z.read('xl/worksheets/sheet4.xml').decode('utf-8')
        modified_sheets['xl/worksheets/sheet4.xml'] = fill_sheet_xml(sheet4, goods_rows)

    # 3. 海事船岸活动 → sheet10
    pc_rows = build_port_call_rows(records)
    if pc_rows:
        with zipfile.ZipFile(template_path) as z:
            sheet10 = z.read('xl/worksheets/sheet10.xml').decode('utf-8')
        modified_sheets['xl/worksheets/sheet10.xml'] = fill_sheet_xml(sheet10, pc_rows)

    # 4. 前十港 → sheet6
    top10_rows = build_top10_rows(records)
    if top10_rows:
        with zipfile.ZipFile(template_path) as z:
            sheet6 = z.read('xl/worksheets/sheet6.xml').decode('utf-8')
        modified_sheets['xl/worksheets/sheet6.xml'] = fill_sheet_xml(sheet6, top10_rows)

    # 5. 复制整个 zip
    with zipfile.ZipFile(template_path, 'r') as zin:
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename in modified_sheets:
                    zout.writestr(item, modified_sheets[item.filename].encode('utf-8'))
                else:
                    zout.writestr(item, zin.read(item.filename))

    # 6. 如果是 xlsx 模式, 改造文件:
    #    - 移除 vbaProject.bin
    #    - 修改 [Content_Types].xml (去掉 vbaProject 声明)
    #    - 修改 workbook.xml (去掉 vbaProject 引用)
    #    - 修改 _rels/workbook.xml.rels (去掉 vbaProject relationship)
    if output_format == 'xlsx':
        convert_to_xlsx(output_path)


def convert_to_xlsx(xlsm_path):
    """把 xlsm 文件就地改造成 xlsx (去掉宏, Excel 不再警告)

    保留: sharedStrings, dataValidations, 所有 sheet, customXml
    移除: vbaProject.bin + 相关引用
    """
    import shutil
    tmp = xlsm_path + '.tmp.zip'
    with zipfile.ZipFile(xlsm_path, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                name = item.filename
                # 跳过 vbaProject
                if 'vbaProject' in name:
                    continue
                data = zin.read(name)

                if name == '[Content_Types].xml':
                    # 去掉 vbaProject 声明
                    text = data.decode('utf-8')
                    text = re.sub(
                        r'<Default Extension="bin" ContentType="application/vnd\.ms-office\.vbaProject"/>',
                        '',
                        text,
                    )
                    # 改 macroEnabled.main -> main
                    text = text.replace(
                        'application/vnd.ms-excel.sheet.macroEnabled.main+xml',
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml',
                    )
                    data = text.encode('utf-8')

                elif name == 'xl/_rels/workbook.xml.rels':
                    # 去掉 vbaProject relationship
                    text = data.decode('utf-8')
                    text = re.sub(
                        r'<Relationship Id="[^"]*" Type="[^"]*/vbaProject" Target="vbaProject\.bin"/>',
                        '',
                        text,
                    )
                    data = text.encode('utf-8')

                zout.writestr(item, data)

    shutil.move(tmp, xlsm_path)


# ============================================================
# CLI
# ============================================================

def main():
    import argparse
    ap = argparse.ArgumentParser(description="单证录入核心 v3 (zipfile 直操作)")
    ap.add_argument("crew", nargs="?", default=DEFAULT_CREW, help="IMO Crew List xlsx 路径（可省略以仅处理 Port of Call）")
    ap.add_argument("port", nargs="?", default=DEFAULT_POC, help="Port of Call xlsx 路径")
    ap.add_argument("output", nargs="?", default=DEFAULT_OUTPUT, help="输出文件路径")
    ap.add_argument("--format", choices=['xlsx', 'xlsm'], default='xlsx',
                    help="输出格式: xlsx (推荐, Excel 不弹警告) 或 xlsm (保留宏, 可能弹警告)")
    args = ap.parse_args()

    if args.format == 'xlsm' and not args.output.endswith('.xlsm'):
        args.output = args.output.replace('.xlsx', '.xlsm')

    if not os.path.exists(TEMPLATE_XLSM):
        print(f"❌ 模板不存在: {TEMPLATE_XLSM}")
        sys.exit(1)
    if not os.path.exists(args.port):
        print(f"❌ 港口文件不存在: {args.port}")
        sys.exit(1)

    has_crew = args.crew and os.path.exists(args.crew)
    if not has_crew:
        print("⚠️  未提供船员文件 — 船员名单/物品清单将不填充")
        crews = []
    else:
        crews = parse_crew_xlsx(args.crew)
        print(f"  解析船员 {len(crews)} 人")

    records = parse_poc_xlsx(args.port)
    print(f"  解析港口 {len(records)} 条")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    process(TEMPLATE_XLSM, args.output, crews, records, output_format=args.format)
    print(f"\n🎉 输出 ({args.format}): {args.output}  ({os.path.getsize(args.output):,} bytes)")


if __name__ == "__main__":
    main()
