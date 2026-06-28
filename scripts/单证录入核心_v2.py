#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
单证录入核心脚本 v2 — 基于 .xlsm 17-sheet 海事标准模板
============================================================
将 IMOCREW LIST (XLSX) + Port of Call (XLSX) 一键转换为
完整的海事局标准录入格式 (.xlsm)。

输出包含 17 个工作表（按海事局标准）：
  1. 船上非旅客人员清单  ✅ 自动填充
  2. 旅客清单
  3. 供退物料清单
  4. 船上非旅客人员物品清单  ✅ 自动填充
  5. 船用物品清单
  6. 前十港信息  ✅ 自动填充
  7. 危险品信息
  8. 船舶证书信息
  9. 压舱水详细信息
 10. 海事船岸活动信息  ✅ 自动填充
 11. 沿海空箱信息
 12-15. 压舱水报告单/装载/更换/排放信息
 16. 随船人员清单
 17. 参数（下拉验证数据源）

用法:
    python3 scripts/单证录入核心_v2.py
    # 或
    python3 scripts/单证录入核心_v2.py <crew.xlsx> <port.xlsx> [output.xlsm]
"""

import json
import re
import sys
import openpyxl
from datetime import datetime
import random
import os
import shutil


WORKDIR = "/Users/qianlingbo/单证录入工作区"
INPUT_DIR = os.path.join(WORKDIR, "input")
OUTPUT_DIR = os.path.join(WORKDIR, "output")
TEMPLATES_DIR = os.path.join(WORKDIR, "templates")
SCRIPTS_DIR = os.path.join(WORKDIR, "scripts")

# 默认输入输出
DEFAULT_CREW = os.path.join(INPUT_DIR, "CREW_LIST.xlsx")
DEFAULT_POC = os.path.join(INPUT_DIR, "PORT_OF_CALL.xlsx")
TEMPLATE_XLSM = os.path.join(TEMPLATES_DIR, "标准格式.xlsm")
DEFAULT_OUTPUT = os.path.join(OUTPUT_DIR, "单证录入标准格式.xlsm")


# 字段映射 ============================================================

NAT_MAP = {
    "CHINESE": "CN", "中国": "CN",
    "VIETNAM": "VN", "越南": "VN",
    "MYANMAR": "MM", "缅甸": "MM",
    "INDONESIA": "ID", "印度尼西亚": "ID",
    "PANAMA": "PA", "巴拿马": "PA",
    "INDIA": "IN", "印度": "IN",
    "PHILIPPINES": "PH", "菲律宾": "PH",
}

NAT_CN = {
    "CHINESE": "中国", "中国": "中国",
    "VIETNAM": "越南", "越南": "越南",
    "MYANMAR": "缅甸", "缅甸": "缅甸",
    "INDONESIA": "印度尼西亚", "印度尼西亚": "印度尼西亚",
    "PANAMA": "巴拿马", "巴拿马": "巴拿马",
    "INDIA": "印度", "印度": "印度",
    "PHILIPPINES": "菲律宾", "菲律宾": "菲律宾",
}

RANK_MAP = {
    "MASTER": "51", "C/O": "52", "2/O": "53", "3/O": "54",
    "OS": "55", "BOSUN": "56",
    "C/E": "61", "2/E": "62", "3/E": "63",
    "ETR": "65", "FTR": "65", "OLR": "55", "AB": "55",
    "C/CK": "55", "CHIEF COOK": "55",
}

# 停靠港口 -> 港口代码
PORT_MAP = {
    "ZHOUSHAN": "CNZOS",       # 舟山
    "ZHANGJIAGANG": "CNZJG",   # 张家港
    "CHANGZHOU": "CNCZX",      # 常州
    "TIANJIN": "CNTXG",        # 天津新港
    "LANSHAN": "CNLSN",        # 岚山
    "BAYUQUAN": "CNBYQ",       # 鲅鱼圈
    "CAOFEIDIAN": "CNCFD",     # 曹妃甸
    "LIANYUNGANG": "CNLYG",    # 连云港
    "NINGDE": "CNNDS",         # 宁德沙埕
    "MOROWALI": "IDMOR",       # 莫罗瓦利
    "POMALAA": "IDPUM",        # 波马拉
    "OBI ISLAND": "IDOBI",     # 奥比岛
    "CHENJIAGANG": None,       # 陈家港 → 找不到，保留空白
    "KENDARI": None,           # 肯达里 → 找不到，保留空白
    "OPEN SEA": "OPSEA",       # 开放海域 → 公海
}

# 登船地点 -> 港口代码
SIGNOFF_PORT_MAP = {
    "NINGDE": "CNNDS",
    "LANSHAN": "CNLSN",
    "CAOFEIDIAN": "CNCFD",
    "BAYUQUAN": "CNBYQ",
    "ZHANGJIAGANG": "CNZJG",
    "CHANGZHOU": "CNCZX",
    "TIANJIN": "CNTXG",
    "LIANYUNGANG": "CNLYG",
}


# 数据解析 ============================================================

def parse_crew_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["ARR"]
    rows = list(ws.iter_rows(values_only=True))
    crews = []
    i = 8
    while i < len(rows):
        row = rows[i]
        if not any(cell is not None for cell in row):
            i += 1
            continue
        no = row[0]
        if not isinstance(no, int):
            i += 1
            continue
        name_en = (row[1] or "").strip() if row[1] else ""
        sex = (row[2] or "").strip() if row[2] else ""
        rank = (row[3] or "").strip() if row[3] else ""
        dob = row[4]
        nat_en = (row[5] or "").strip() if row[5] else ""
        sob = row[6]
        seaman_book = (row[7] or "").strip() if row[7] else ""
        seaman_expiry = row[8]
        passport = (row[9] or "").strip() if row[9] else ""
        passport_expiry = row[10]

        name_cn = ""
        signon_port = ""
        if i + 1 < len(rows):
            nrow = rows[i + 1]
            if nrow[1]:
                name_cn = (nrow[1] or "").strip()
            if nrow[6]:
                signon_port = (nrow[6] or "").strip()

        crews.append({
            "no": no, "name_en": name_en, "name_cn": name_cn,
            "sex": sex, "rank": rank, "dob": dob, "nat_en": nat_en,
            "signon_date": sob, "seaman_book": seaman_book,
            "seaman_expiry": seaman_expiry, "passport": passport,
            "passport_expiry": passport_expiry, "signon_port": signon_port,
        })
        i += 2
    return crews


def parse_poc_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    ports = []
    for row in rows[3:]:
        no = row[0]
        if not isinstance(no, int):
            continue
        purpose = (row[1] or "").strip() if row[1] else ""
        port = (row[2] or "").strip() if row[2] else ""
        country = (row[3] or "").strip() if row[3] else ""
        arr = (row[4] or "").strip() if row[4] else ""
        dep = (row[5] or "").strip() if row[5] else ""
        ssl = row[6]
        psl = row[7]
        ports.append({
            "no": no, "purpose": purpose, "port": port, "country": country,
            "arr": arr, "dep": dep, "ssl": ssl, "psl": psl,
        })
    return ports


# 转换函数 ============================================================

def fmt_date_yyyymmdd(dt):
    if isinstance(dt, datetime):
        return dt.strftime("%Y%m%d")
    return ""


def map_nationality(nat_en):
    return NAT_MAP.get(nat_en.upper(), "")


def map_birth_place(nat_en):
    """出生地点：船员国籍的中文（按 SKILL 规则）"""
    return NAT_CN.get(nat_en.upper(), "")


def map_rank(rank_en):
    if not rank_en:
        return ""
    r = rank_en.strip().upper()
    if r in RANK_MAP:
        return RANK_MAP[r]
    for key, val in RANK_MAP.items():
        if key in r or r in key:
            return val
    return "55"


def map_sex(sex):
    if not sex:
        return ""
    s = sex.strip().upper()
    if s.startswith("M"): return "1"
    if s.startswith("F"): return "2"
    return "1"


def is_chinese(nat_en):
    return nat_en and nat_en.upper() in ("CHINESE", "中国")


def normalize_date(s):
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    txt = re.sub(r"\s+", "", str(s)).strip()
    m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})(\d{1,2}):(\d{1,2}):(\d{1,2})", txt)
    if not m:
        m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", txt)
        if not m:
            return None
        y, mo, d = m.groups()
        return datetime(int(y), int(mo), int(d), 0, 0, 0)
    y, mo, d, h, mi, se = m.groups()
    try:
        return datetime(int(y), int(mo), int(d), int(h), int(mi), int(se))
    except ValueError:
        return None


def map_country_code(country):
    return country.upper() if country else ""


def map_port_code(port_en):
    if not port_en:
        return None
    key = port_en.strip().upper()
    if key in PORT_MAP:
        return PORT_MAP[key]
    for k, v in PORT_MAP.items():
        if k in key or key in k:
            return v
    return None


def map_signon_port_code(port_en):
    if not port_en:
        return None
    key = port_en.strip().upper()
    if key in SIGNOFF_PORT_MAP:
        return SIGNOFF_PORT_MAP[key]
    for k, v in SIGNOFF_PORT_MAP.items():
        if k in key or key in k:
            return v
    return None


# 生成 ============================================================

def fill_crew_list(ws, crews):
    """填充『船上非旅客人员清单』"""
    for i, crew in enumerate(crews):
        row_idx = 3 + i
        if is_chinese(crew["nat_en"]):
            name = crew["name_cn"]
            cert_type = "17"
            cert_no = crew["seaman_book"]
        else:
            name = crew["name_en"].upper()
            cert_type = "14"
            cert_no = crew["passport"]

        sex = map_sex(crew["sex"])
        rank = map_rank(crew["rank"])
        nat_code = map_nationality(crew["nat_en"])
        dob = fmt_date_yyyymmdd(crew["dob"])
        birth_place = map_birth_place(crew["nat_en"])
        signon_date = fmt_date_yyyymmdd(crew["signon_date"])
        signon_port = map_signon_port_code(crew["signon_port"]) or ""

        ws.cell(row=row_idx, column=1, value=crew["no"])
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=sex)
        ws.cell(row=row_idx, column=4, value=rank)
        ws.cell(row=row_idx, column=5, value=nat_code)
        ws.cell(row=row_idx, column=6, value=dob)
        ws.cell(row=row_idx, column=7, value=birth_place)
        ws.cell(row=row_idx, column=8, value=cert_type)
        ws.cell(row=row_idx, column=9, value=cert_no)
        # 10-13 空白
        ws.cell(row=row_idx, column=14, value=signon_date)
        ws.cell(row=row_idx, column=15, value=signon_port)


def fill_goods_list(ws, crews):
    """填充『船上非旅客人员物品清单』"""
    for i, crew in enumerate(crews):
        row_idx = 3 + i
        if is_chinese(crew["nat_en"]):
            cert_type = "17"
            cert_no = crew["seaman_book"]
        else:
            cert_type = "14"
            cert_no = crew["passport"]

        ws.cell(row=row_idx, column=1, value=crew["no"])
        ws.cell(row=row_idx, column=2, value=cert_type)
        ws.cell(row=row_idx, column=3, value=cert_no)
        # 物品类型 0100-计算机 (D列)
        ws.cell(row=row_idx, column=4, value="0100")
        # 物品名称
        ws.cell(row=row_idx, column=5, value="计算机")
        # 物品数量
        ws.cell(row=row_idx, column=6, value=1)


def fill_port_call(ws, poc_list):
    """填充『海事船岸活动信息』"""
    for i, p in enumerate(poc_list):
        row_idx = 3 + i
        ws.cell(row=row_idx, column=1, value=i + 1)

        arr_dt = normalize_date(p["arr"])
        if arr_dt:
            hr = random.randint(0, 11)
            mi = random.randint(0, 59)
            arr_str = arr_dt.strftime("%Y/%m/%d") + f" {hr:02d}:{mi:02d}:00"
        else:
            arr_str = ""
        ws.cell(row=row_idx, column=2, value=arr_str)

        dep_dt = normalize_date(p["dep"])
        if dep_dt:
            hr = random.randint(12, 23)
            mi = random.randint(0, 59)
            dep_str = dep_dt.strftime("%Y/%m/%d") + f" {hr:02d}:{mi:02d}:00"
        else:
            dep_str = ""
        ws.cell(row=row_idx, column=3, value=dep_str)

        country = map_country_code(p["country"])
        ws.cell(row=row_idx, column=4, value=country)
        ws.cell(row=row_idx, column=5, value="1-1级")

        port_code = map_port_code(p["port"])
        ws.cell(row=row_idx, column=7, value=port_code or "")
        ws.cell(row=row_idx, column=8, value="1-1级")


def fill_top10_ports(ws, poc_list):
    """填充『前十港信息』 — 用 Port of Call 数据填前 10 条
    列: A=序号 B=进港时间 C=离港时间 D=航次号 E=国家/地区 F=停靠港口 G=目的...
    """
    for i, p in enumerate(poc_list[:10]):
        row_idx = 3 + i
        ws.cell(row=row_idx, column=1, value=i + 1)

        arr_dt = normalize_date(p["arr"])
        if arr_dt:
            hr = random.randint(0, 11)
            mi = random.randint(0, 59)
            arr_str = arr_dt.strftime("%Y/%m/%d") + f" {hr:02d}:{mi:02d}:00"
        else:
            arr_str = ""
        ws.cell(row=row_idx, column=2, value=arr_str)

        dep_dt = normalize_date(p["dep"])
        if dep_dt:
            hr = random.randint(12, 23)
            mi = random.randint(0, 59)
            dep_str = dep_dt.strftime("%Y/%m/%d") + f" {hr:02d}:{mi:02d}:00"
        else:
            dep_str = ""
        ws.cell(row=row_idx, column=3, value=dep_str)

        # 航次号 - 留空
        country = map_country_code(p["country"])
        ws.cell(row=row_idx, column=5, value=country)
        port_code = map_port_code(p["port"])
        ws.cell(row=row_idx, column=6, value=port_code or "")


def main():
    import argparse
    ap = argparse.ArgumentParser(description="单证录入核心 v2 (.xlsm 17-sheet 模板)")
    ap.add_argument("crew", nargs="?", default=DEFAULT_CREW, help="IMO Crew List xlsx 路径")
    ap.add_argument("port", nargs="?", default=DEFAULT_POC, help="Port of Call xlsx 路径")
    ap.add_argument("output", nargs="?", default=DEFAULT_OUTPUT, help="输出 .xlsm 路径")
    args = ap.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if not os.path.exists(TEMPLATE_XLSM):
        print(f"❌ 模板不存在: {TEMPLATE_XLSM}")
        sys.exit(1)
    if not os.path.exists(args.crew):
        print(f"❌ 船员文件不存在: {args.crew}")
        sys.exit(1)
    if not os.path.exists(args.port):
        print(f"❌ 港口文件不存在: {args.port}")
        sys.exit(1)

    # 1. 复制模板作为基础
    shutil.copy(TEMPLATE_XLSM, args.output)
    print(f"📋 复制模板 → {args.output}")

    # 2. 解析输入数据
    crews = parse_crew_xlsx(args.crew)
    poc_list = parse_poc_xlsx(args.port)
    print(f"  解析船员 {len(crews)} 人, 港口记录 {len(poc_list)} 条")

    # 3. 打开模板并填充
    wb = openpyxl.load_workbook(args.output)

    # 船员名单
    if "船上非旅客人员清单" in wb.sheetnames:
        fill_crew_list(wb["船上非旅客人员清单"], crews)
        print("  ✅ 船上非旅客人员清单")

    # 物品清单
    if "船上非旅客人员物品清单" in wb.sheetnames:
        fill_goods_list(wb["船上非旅客人员物品清单"], crews)
        print("  ✅ 船上非旅客人员物品清单")

    # 海事船岸活动
    if "海事船岸活动信息" in wb.sheetnames:
        fill_port_call(wb["海事船岸活动信息"], poc_list)
        print("  ✅ 海事船岸活动信息")

    # 前十港信息
    if "前十港信息" in wb.sheetnames:
        fill_top10_ports(wb["前十港信息"], poc_list)
        print("  ✅ 前十港信息")

    wb.save(args.output)
    print(f"\n🎉 全部完成: {args.output}")
    print(f"   文件大小: {os.path.getsize(args.output):,} bytes")


if __name__ == "__main__":
    main()
